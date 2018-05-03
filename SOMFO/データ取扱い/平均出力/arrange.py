import numpy
import glob
import sys;
import openpyxl as px
from openpyxl.styles import Alignment
import os;
from openpyxl.styles import Border, Side
import string
import numpy as np
import glob 
import re



def numericalSort(value):
    numbers = re.compile(r'(\d+)')
    parts = numbers.split(value)
    parts[1::2] = map(int, parts[1::2])
    return parts
def getAllFileName_sort(str_,target):
	return sorted(glob.iglob(str_+'/**/' + target, recursive=True), key=numericalSort);

def getAllFileName(dir,target):
	return glob.iglob(dir+'/**/' + target, recursive=True);

def setAverage(targetdir,targetProblem):
	FILES = getAllFileName(targetdir,targetProblem);
	print(targetdir + "	" + targetProblem)
	for file in FILES:
#		print(file);
		output = file.split("\\")[1]
		output_every = output.split("amount")[0];
		output_amount = "amount"+output.split("amount")[1];
		#print(output_every + "	" + output_amount);
		DATAFILES = getAllFileName_sort(file,"BestFUN");
		for datadir in DATAFILES:
			getAveAndSTD(datadir);
#			print("delecate");

# return touple, first element is average and second element is std
def getAveAndSTD(targetdir):
	FILES_ = glob.glob(targetdir+"/*.dat");
	#print(targetdir);
	data_ = np.array([]);
	for f in FILES_:
		if ("average" in f):
			continue;
		with open(f,"r") as fin:
			data = fin.readlines();
			last = data[len(data)-1].replace("\n","").split("	")[1];
#			print(last);
			data_ = np.append(data_,last);
	data_ = data_.astype(np.float32);
	return (np.mean(data_),np.std(data_));

def getOrderHashMap(targetdir):
	FILES = getAllFileName(targetdir,"BestFUN");
	EveryHashmap = {};
	AmountHashmap = {};
	EVERYList = [];
	AmountList = [];
	d = 0;
	for file in FILES:
		#if d == 0:
			#d = d+1;
		output = file.split("\\")[1];
		output_every = output.split("amount")[0];
		output_amount = "amount"+output.split("amount")[1];
		if (output_every not in EVERYList ):
			EVERYList.append(output_every);
		if (output_amount not in AmountList ):
			AmountList.append(output_amount);

	AmountList = sorted(AmountList, key=numericalSort);
	EVERYList = sorted(EVERYList, key=numericalSort);
#	print(AmountList)
#	print(EVERYList)
	counter = 0;
	for d in EVERYList:
		EveryHashmap[d] = counter;
		counter  = counter+1;
	counter = 0;
	for d in AmountList:
		AmountHashmap[d] = counter;
		counter  = counter+1;
	return (EveryHashmap,AmountHashmap)

def getOrderHashMap_only_merge(targetdir):
	FILES = getAllFileName(targetdir,"BestFUN");
	EveryHashmap = {};
	AmountHashmap = {};
	EVERYList = [];
	AmountList = [];
	d = 0;
	for file in FILES:
		#if d == 0:
			#d = d+1;
		output = file.split("\\")[1];
		output_every = output.split("amount")[0];
		if (output_every not in EVERYList ):
			EVERYList.append(output_every);
		
	AmountList = sorted(AmountList, key=numericalSort);
	EVERYList = sorted(EVERYList, key=numericalSort);
#	print(AmountList)
#	print(EVERYList)
	counter = 0;
	for d in EVERYList:
		EveryHashmap[d] = counter;
		counter  = counter+1;
	counter = 0;
	for d in AmountList:
		AmountHashmap[d] = counter;
		counter  = counter+1;
	return (EveryHashmap,AmountHashmap)


def getList(targetdir,targetProblem):
	FILES = getAllFileName(targetdir,targetProblem);
	EveryHashmap = {};
	AmountHashmap = {};
	EVERYList = [];
	AmountList = [];
	d = 0;
	for file in FILES:
		#if d == 0:
			#d = d+1;
		output = file.split("\\")[1];
		output_every = output.split("amount")[0];
#		output_amount = "amount"+output.split("amount")[1];
		if (output_every not in EVERYList ):
			EVERYList.append(output_every);
	#	if (output_amount not in AmountList ):
	#		AmountList.append(output_amount);
	AmountList = sorted(AmountList, key=numericalSort);
	EVERYList = sorted(EVERYList, key=numericalSort);
	return (EVERYList,AmountList)


def outputfile_without_merge(data_algorithm):
	PROBLEMSET = ["CIHS","CIMS","CILS","PIHS","PIMS","PILS","NIHS","NIMS","NILS"]
	Algorithm = data_algorithm
#	for prob in PROBLEMSET:
#		getOrderHashMap("Random",prob)
#		setAverage("Random",prob);
	problemname = "CIHS"
	d = getOrderHashMap(Algorithm);
	EveryHashmap = d[0];
	AmountHashmap = d[1];
	ProblemHashmap = {}

	for i in range(0,len(PROBLEMSET)):
			ProblemHashmap[PROBLEMSET[i]] = i;
	Lits = getList(Algorithm,"CIHS");
	EveryList = Lits[0];
	AmounList = Lits[1];

	wb = px.Workbook();	
	ws = wb.get_sheet_by_name("Sheet");
	Alpharbet = [];
	d = list(string.ascii_uppercase);
	
	for alp in d:
		Alpharbet.append(alp);
	
	for alp_one in d:
		for alp_two in d:
			Alpharbet.append(alp_one+alp_two);
	

	counter_first=0;
	counter_second=0;
#	for d in
	ws.cell(row=1,column=1).value = Algorithm

##setting writting
###############################################################################################################################################################################################
	PRLBLEMSCOUNT = 13;
	for amountli in AmountHashmap:
		counter_first = counter_first+1
		counter_second=0;
		for everyli in EveryList:
			counter_second = counter_second+1;
			ws.merge_cells(Alpharbet[4*(counter_second-1)+1]+str(PRLBLEMSCOUNT*(counter_first-1)+1)+":"+Alpharbet[4*(counter_second)]+str(PRLBLEMSCOUNT*(counter_first-1)+1));
			ws.merge_cells(Alpharbet[4*(counter_second-1)+1]+str(PRLBLEMSCOUNT*(counter_first-1)+2)+":"+Alpharbet[4*(counter_second-1)+2]+str(PRLBLEMSCOUNT*(counter_first-1)+2));
			ws.merge_cells(Alpharbet[4*(counter_second-1)+3]+str(PRLBLEMSCOUNT*(counter_first-1)+2)+":"+Alpharbet[4*(counter_second-1)+4]+str(PRLBLEMSCOUNT*(counter_first-1)+2));
			
			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+2,column =4*(counter_second-1)+2).value ="Task1"
			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+2,column =4*(counter_second-1)+4 ).value = "Task2"
			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+2).value ="ave"
			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+3).value ="std"
			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+4).value ="ave"
			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+5 ).value = "std"
			ws[Alpharbet[4*(counter_second-1)+2]+str(PRLBLEMSCOUNT*(counter_first-1)+3)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
			ws[Alpharbet[4*(counter_second-1)+4]+str(PRLBLEMSCOUNT*(counter_first-1)+3)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
			ws[Alpharbet[4*(counter_second-1)+2]+str(PRLBLEMSCOUNT*(counter_first-1)+2)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
			ws[Alpharbet[4*(counter_second-1)+4]+str(PRLBLEMSCOUNT*(counter_first-1)+2)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
			ws[Alpharbet[4*(counter_second-1)+1]+str(PRLBLEMSCOUNT*(counter_first-1)+1)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')			


			ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+1,column =4*(counter_second-1)+2 ).value = everyli+amountli
	for i in range(0,counter_first):
		pcoun = 0;
		for j in range(0,len(PROBLEMSET)):
			ws.cell(row = i*(counter_first+3)+ j+4,column =1).value =PROBLEMSET[j]
###############################################################################################################################################################################################
	FILES_DATA = getAllFileName(Algorithm,"BestFUN");
	for file in FILES_DATA:
		filesplitname = file.split("\\");
		everydata = filesplitname[1].split("amount")[0]
		amountdata = "amount"+filesplitname[1].split("amount")[1]

		print(file)
		print(everydata + "	" + amountdata);
		ave_std_data = getAveAndSTD(file);
		ave_data = ave_std_data[0];
		std_data = ave_std_data[1];
		PROBLEMDATA = filesplitname[2]
		TaskNumber = filesplitname[len(filesplitname)-2].split("Task")[1]; 
		print(TaskNumber)
		print(str(ave_data)+"	" + str(std_data))
		print(PROBLEMDATA);
		every_ = EveryHashmap[everydata];
		amount_ = AmountHashmap[amountdata];
		print(str(every_)+"	" + str(amount_))	
		ws.cell(row = PRLBLEMSCOUNT*(amount_)+3 + ProblemHashmap[PROBLEMDATA] +1,column =4*(every_)+2*int(TaskNumber)).value =ave_data;
		ws.cell(row = PRLBLEMSCOUNT*(amount_)+3 + ProblemHashmap[PROBLEMDATA] +1 ,column =4*(every_)+2*int(TaskNumber)+1).value = std_data;
		
	wb.save(Algorithm+".xlsx");



def outputfile_only_merge(data_algorithm):
	PROBLEMSET = ["CIHS","CIMS","CILS","PIHS","PIMS","PILS","NIHS","NIMS","NILS"]
	Algorithm = data_algorithm
#	for prob in PROBLEMSET:
#		getOrderHashMap("Random",prob)
#		setAverage("Random",prob);
	problemname = "CIHS"
	d = getOrderHashMap_only_merge(Algorithm);
	EveryHashmap = d[0];
	AmountHashmap = d[1];
	ProblemHashmap = {}

	for i in range(0,len(PROBLEMSET)):
			ProblemHashmap[PROBLEMSET[i]] = i;
	Lits = getList(Algorithm,"CIHS");
	EveryList = Lits[0];
	AmounList = Lits[1];

	wb = px.Workbook();	
	ws = wb.get_sheet_by_name("Sheet");
	Alpharbet = [];
	d = list(string.ascii_uppercase);
	
	for alp in d:
		Alpharbet.append(alp);
	
	for alp_one in d:
		for alp_two in d:
			Alpharbet.append(alp_one+alp_two);
	

	counter_first=0;
	counter_second=0;
#	for d in
	ws.cell(row=1,column=1).value = Algorithm

##setting writting
###############################################################################################################################################################################################
	PRLBLEMSCOUNT = 13;
	counter_first = counter_first+1
	counter_second=0;
	for everyli in EveryList:
		counter_second = counter_second+1;
		ws.merge_cells(Alpharbet[4*(counter_second-1)+1]+str(PRLBLEMSCOUNT*(counter_first-1)+1)+":"+Alpharbet[4*(counter_second)]+str(PRLBLEMSCOUNT*(counter_first-1)+1));
		ws.merge_cells(Alpharbet[4*(counter_second-1)+1]+str(PRLBLEMSCOUNT*(counter_first-1)+2)+":"+Alpharbet[4*(counter_second-1)+2]+str(PRLBLEMSCOUNT*(counter_first-1)+2));
		ws.merge_cells(Alpharbet[4*(counter_second-1)+3]+str(PRLBLEMSCOUNT*(counter_first-1)+2)+":"+Alpharbet[4*(counter_second-1)+4]+str(PRLBLEMSCOUNT*(counter_first-1)+2));
		
		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+2,column =4*(counter_second-1)+2).value ="Task1"
		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+2,column =4*(counter_second-1)+4 ).value = "Task2"
		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+2).value ="ave"
		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+3).value ="std"
		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+4).value ="ave"
		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+3,column =4*(counter_second-1)+5 ).value = "std"
		ws[Alpharbet[4*(counter_second-1)+2]+str(PRLBLEMSCOUNT*(counter_first-1)+3)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
		ws[Alpharbet[4*(counter_second-1)+4]+str(PRLBLEMSCOUNT*(counter_first-1)+3)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
		ws[Alpharbet[4*(counter_second-1)+2]+str(PRLBLEMSCOUNT*(counter_first-1)+2)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
		ws[Alpharbet[4*(counter_second-1)+4]+str(PRLBLEMSCOUNT*(counter_first-1)+2)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
		ws[Alpharbet[4*(counter_second-1)+1]+str(PRLBLEMSCOUNT*(counter_first-1)+1)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')			


		ws.cell(row = PRLBLEMSCOUNT*(counter_first-1)+1,column =4*(counter_second-1)+2 ).value = everyli
	for i in range(0,counter_first):
		pcoun = 0;
		for j in range(0,len(PROBLEMSET)):
			ws.cell(row = i*(counter_first+3)+ j+4,column =1).value =PROBLEMSET[j]
###############################################################################################################################################################################################
	FILES_DATA = getAllFileName(Algorithm,"BestFUN");
	for file in FILES_DATA:
		filesplitname = file.split("\\");
		everydata = filesplitname[1].split("amount")[0]
#		amountdata = "amount"+filesplitname[1].split("amount")[1]

		print(file)
		print(everydata);
		ave_std_data = getAveAndSTD(file);
		ave_data = ave_std_data[0];
		std_data = ave_std_data[1];
		PROBLEMDATA = filesplitname[2]
		TaskNumber = filesplitname[len(filesplitname)-2].split("Task")[1]; 
		print(TaskNumber)
		print(str(ave_data)+"	" + str(std_data))
		print(PROBLEMDATA);
		every_ = EveryHashmap[everydata];
	#	amount_ = AmountHashmap[amountdata];
		amount_ = 0;
		print(str(every_)+"	" + str(amount_))	
		ws.cell(row = PRLBLEMSCOUNT*(amount_)+3 + ProblemHashmap[PROBLEMDATA] +1,column =4*(every_)+2*int(TaskNumber)).value =ave_data;
		ws.cell(row = PRLBLEMSCOUNT*(amount_)+3 + ProblemHashmap[PROBLEMDATA] +1 ,column =4*(every_)+2*int(TaskNumber)+1).value = std_data;
		
	wb.save(Algorithm+".xlsx");



if __name__ == "__main__":
	print("start");
	data = sys.argv[1];
	if (data == "Merge" or data == "merge"):
		outputfile_only_merge(data);
	else:
		outputfile_without_merge(data);

	
	print("end")
